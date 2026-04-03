from __future__ import annotations

import calendar
import json
import os
import re
import difflib
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from sqlalchemy import text

#from db import engine
from db_cloud import engine


from forecast import compute_forecast_from_db

from etl.bulls import read_bulls_txt, load_bulls_to_db
from etl.calvings_births import read_calvings_excel, load_calvings_to_db
from etl.disposals import read_disposals_excel, load_disposals_to_db
from etl.dryoff import read_dryoff_excel, load_dryoff_to_db
from etl.inseminations import read_inseminations_excel, clean_inseminations, load_inseminations_to_db

import model_params as mp


# ============================================================
# Константы для UI
# ============================================================
INDICATORS = [
    "Дойные коровы",
    "Сухостойные коровы",
    "Тёлки 0–3 мес",
    "Бычки 0–2 мес",
    "Тёлки 3–8 мес",
    "Тёлки ≥9 мес",
    "Нетели",
    "Ожидаемый отёл, всего",
    "Ожидаемый отёл, из них коров",
    "Ожидаемый отёл, из них нетелей",
    "Ожидаемые бычки (условно)",
    "Ожидаемые тёлочки (условно)",
]

OVERFLOW_COLS = [
    "К реализации: коровы",
    "К реализации: тёлки",
    "К реализации: нетели",
    "Переполнение: Дойные коровы",
    "Переполнение: Сухостойные коровы",
    "Переполнение: Тёлки 0–3 мес",
    "Переполнение: Тёлки 3–8 мес",
    "Переполнение: Тёлки 9–24 мес",
    "Переполнение: Нетели",
]

# По твоей просьбе: в UI показываем ТОЛЬКО переполнение по группам
OVERFLOW_GROUP_COLS = [c for c in OVERFLOW_COLS if c.startswith("Переполнение:")]


# ============================================================
# Общие helpers
# ============================================================
def month_end(y: int, m: int) -> date:
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last)


def iter_month_ends(y1: int, m1: int, y2: int, m2: int) -> List[date]:
    out: List[date] = []
    y, m = y1, m1
    while (y, m) <= (y2, m2):
        out.append(month_end(y, m))
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    return out


def _fmt_cell(x: Any) -> Any:
    try:
        if x is None:
            return ""
        f = float(x)
        if pd.isna(f):
            return ""
        s = f"{f:.1f}".replace(",", ".")
        s = s.rstrip("0").rstrip(".")
        return s
    except Exception:
        return x


def ensure_month_col(df: pd.DataFrame, month_labels: list[str] | None = None) -> pd.DataFrame:
    if "Месяц" in df.columns:
        return df

    rename_map = {}
    for cand in ("month", "Month", "MONTH", "Дата", "date", "Date"):
        if cand in df.columns:
            rename_map[cand] = "Месяц"
            break
    if rename_map:
        df = df.rename(columns=rename_map)
        if "Месяц" in df.columns:
            return df

    if month_labels is not None and len(month_labels) >= len(df):
        df = df.copy()
        df.insert(0, "Месяц", month_labels[:len(df)])
        return df

    df = df.copy()
    df.insert(0, "Месяц", [str(i) for i in range(1, len(df) + 1)])
    return df


def get_max_event_date_from_db() -> date:
    q = """
    SELECT
      GREATEST(
        (SELECT MAX(event_date) FROM calvings_births_raw),
        (SELECT MAX(event_date) FROM inseminations_raw),
        (SELECT MAX(event_date) FROM dryoff_raw),
        (SELECT MAX(event_date) FROM disposals_raw)
      ) AS max_date;
    """
    try:
        df = pd.read_sql(q, con=engine)
        v = df.loc[0, "max_date"]
        if pd.isna(v):
            return date.today()
        return pd.to_datetime(v).date()
    except Exception:
        return date.today()


# ============================================================
# Нормализация ключей из compute_forecast_from_db (на случай тире/пробелов)
# ============================================================
def _norm_label(x: Any) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\u00a0", " ").strip()
    s = s.replace("Ё", "Е").replace("ё", "е")
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    s = re.sub(r"\s+", " ", s).strip()

    su = s.upper()
    su = su.replace("ТЁЛОЧК", "ТЁЛК")
    su = su.replace("ТЕЛОЧК", "ТЕЛК")
    return su


def _vals_get(vals: dict, want_key: str, norm_map: dict | None = None) -> Any:
    if not isinstance(vals, dict) or not vals:
        return None

    if want_key in vals:
        return vals.get(want_key)

    if norm_map is None:
        norm_map = {_norm_label(k): v for k, v in vals.items()}

    nk = _norm_label(want_key)
    if nk in norm_map:
        return norm_map.get(nk)

    candidates = [k for k in norm_map.keys() if (nk in k) or (k in nk)]
    if len(candidates) == 1:
        return norm_map.get(candidates[0])

    close = difflib.get_close_matches(nk, list(norm_map.keys()), n=1, cutoff=0.92)
    if close:
        return norm_map.get(close[0])

    return None


# ============================================================
# Excel export (оставляем legacy функцию, но UI её больше не использует)
# ============================================================
def make_excel_bytes(forecast_df: pd.DataFrame, realization_df: pd.DataFrame) -> bytes:
    """
    LEGACY (не удаляем).
    """
    import openpyxl  # noqa: F401
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        forecast_df.to_excel(writer, sheet_name="Прогноз")
        realization_df.to_excel(writer, sheet_name="Реализация")

        wb = writer.book
        fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

        def _apply_positive_rule(ws) -> None:
            if ws.max_row < 2 or ws.max_column < 2:
                return
            header_row = 1
            for col_idx in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{header_row+1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator="greaterThan", formula=["0"], fill=fill),
                )

        if "Реализация" in wb.sheetnames:
            _apply_positive_rule(wb["Реализация"])

    return out.getvalue()


def make_excel_bytes_highlight_months_columns(
    forecast_view: pd.DataFrame,   # индикаторы x месяцы
    overflow_view: pd.DataFrame,   # переполнение-группы x месяцы
    indicator_to_overflow: dict[str, str | None],
) -> bytes:
    """
    Новый Excel под текущий UI:
    - Прогноз 
    - Переполнение 
    Подсветка:
      * Переполнение: любое >0 красным
      * Прогноз: ячейка индикатора красная, если в этом месяце переполнение по группе >0
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        forecast_view.to_excel(writer, sheet_name="Прогноз ")
        overflow_view.to_excel(writer, sheet_name="Переполнение ")

    buf.seek(0)
    wb = load_workbook(buf)

    fill = PatternFill("solid", fgColor="FF0000")
    font = Font(color="FFFFFF", bold=True)
    num_fmt = "0.#"

    def _is_pos(v: Any) -> bool:
        try:
            return float(v) > 0.0
        except Exception:
            return False

    # Переполнение: красим любое >0
    ws_ov = wb["Переполнение "]
    for row in ws_ov.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if _is_pos(cell.value):
                cell.fill = fill
                cell.font = font
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt

    # Прогноз: красим если переполнение >0
    ws_f = wb["Прогноз "]

    # Заголовки месяцев в прогнозе
    months = []
    for c in range(2, ws_f.max_column + 1):
        months.append(str(ws_f.cell(row=1, column=c).value))

    # Индикаторы по строкам
    indicators = []
    for r in range(2, ws_f.max_row + 1):
        indicators.append(str(ws_f.cell(row=r, column=1).value))

    # Заголовки в overflow
    ov_month_to_col = {}
    for c in range(2, ws_ov.max_column + 1):
        ov_month_to_col[str(ws_ov.cell(row=1, column=c).value)] = c

    ov_row_by_name = {}
    for r in range(2, ws_ov.max_row + 1):
        ov_row_by_name[str(ws_ov.cell(row=r, column=1).value)] = r

    for r_idx, ind in enumerate(indicators, start=2):
        ov_name = indicator_to_overflow.get(ind)
        if not ov_name:
            continue
        ov_r = ov_row_by_name.get(ov_name)
        if not ov_r:
            continue

        for c_idx, m in enumerate(months, start=2):
            ov_c = ov_month_to_col.get(m)
            if not ov_c:
                continue
            ov_val = ws_ov.cell(row=ov_r, column=ov_c).value
            if _is_pos(ov_val):
                cell = ws_f.cell(row=r_idx, column=c_idx)
                cell.fill = fill
                cell.font = font
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt
            else:
                cell = ws_f.cell(row=r_idx, column=c_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# Параметры из БД (как у тебя раньше) + DB-cache
# ============================================================
def _norm_id(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).replace("\u00a0", " ").strip()
    if s == "" or s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        s2 = s[:-2]
        if s2.isdigit():
            return s2
    return s


def _norm_event_type(x: Any) -> str:
    if x is None:
        return ""
    v = str(x).replace("\xa0", " ").strip().upper().replace("Ё", "Е")
    if "РОЖ" in v:
        return "РОЖДЕН"
    return v


def _norm_result(x: Any) -> str:
    if x is None:
        return ""
    v = str(x).replace("\xa0", " ").strip().upper().replace("Ё", "Е")
    if v in {"P", "П"}:
        return "P"
    if "PREG" in v:
        return "P"
    if "СТЕЛ" in v or v in {"СТ", "СТ.", "СТ+", "СТЕЛЬНАЯ", "СТЕЛЬН"}:
        return "P"
    return v


def _norm_sex(x: Any) -> str | None:
    if x is None:
        return None
    v = str(x).strip().upper()
    if v in ("F", "Ж", "FEMALE"):
        return "F"
    if v in ("M", "М", "MALE"):
        return "M"
    return None


def _classify_semen_from_bull_type(bt: Any) -> str:
    v = "" if bt is None else str(bt).strip().upper()
    if v == "S" or "SEX" in v:
        return "sex"
    return "trad"


def _merge_asof_safe(
    left: pd.DataFrame,
    right: pd.DataFrame,
    *,
    by: str,
    left_on: str,
    right_on: str,
    direction: str = "backward",
    allow_exact_matches: bool = True,
) -> pd.DataFrame:
    l = left.copy()
    r = right.copy()

    l["_row_id"] = range(len(l))

    l[by] = l[by].astype(str).str.strip()
    r[by] = r[by].astype(str).str.strip()

    l[left_on] = pd.to_datetime(l[left_on], errors="coerce")
    r[right_on] = pd.to_datetime(r[right_on], errors="coerce")

    l = l[(l[by] != "") & l[left_on].notna()].copy()
    r = r[(r[by] != "") & r[right_on].notna()].copy()

    l = l.sort_values([left_on, by], kind="mergesort").reset_index(drop=True)
    r = r.sort_values([right_on, by], kind="mergesort").reset_index(drop=True)

    out = pd.merge_asof(
        l,
        r,
        by=by,
        left_on=left_on,
        right_on=right_on,
        direction=direction,
        allow_exact_matches=allow_exact_matches,
    )

    out = out.sort_values("_row_id", kind="mergesort").drop(columns=["_row_id"]).reset_index(drop=True)
    return out


def _bayes_smooth_share(sex_count: float, total: float, prior_sex: float, prior_weight: float) -> float:
    total = float(total)
    sex_count = float(sex_count)
    prior_sex = float(prior_sex)
    prior_weight = float(prior_weight)
    denom = max(1e-9, total + prior_weight)
    p = (sex_count + prior_weight * prior_sex) / denom
    return float(max(0.0, min(1.0, p)))


def _get_db_signature() -> str:
    q = """
    SELECT
      COALESCE((SELECT MAX(event_date)::text FROM calvings_births_raw), '') AS calv_max,
      COALESCE((SELECT MAX(event_date)::text FROM inseminations_raw), '') AS ins_max,
      COALESCE((SELECT MAX(event_date)::text FROM dryoff_raw), '') AS dry_max,
      COALESCE((SELECT MAX(event_date)::text FROM disposals_raw), '') AS disp_max,

      (SELECT COUNT(*) FROM calvings_births_raw) AS calv_n,
      (SELECT COUNT(*) FROM inseminations_raw) AS ins_n,
      (SELECT COUNT(*) FROM dryoff_raw) AS dry_n,
      (SELECT COUNT(*) FROM disposals_raw) AS disp_n
    ;
    """
    try:
        df = pd.read_sql(q, con=engine)
        r = df.iloc[0].to_dict()
        return f"{r['calv_max']}|{r['ins_max']}|{r['dry_max']}|{r['disp_max']}|{r['calv_n']}|{r['ins_n']}|{r['dry_n']}|{r['disp_n']}"
    except Exception:
        return "no-db"


def _ensure_params_cache_table() -> None:
    q = """
    CREATE TABLE IF NOT EXISTS model_params_cache (
        signature TEXT PRIMARY KEY,
        params_json TEXT NOT NULL,
        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    """
    with engine.connect() as conn:
        conn.execute(text(q))
        conn.commit()


#def _ensure_params_cache_table() -> None:
 #   q = """
 #   """CREATE TABLE IF NOT EXISTS model_params_cache (
  #      signature TEXT PRIMARY KEY,
 #       params_json JSONB NOT NULL,
 #       updated_at TIMESTAMP NOT NULL DEFAULT NOW()
  #  );"""
  #  """
 #   with engine.begin() as conn:
  #      conn.execute(text(q))"""


def _load_params_from_db_cache(sig: str) -> Optional[Dict[str, Any]]:
    _ensure_params_cache_table()
    q = "SELECT params_json FROM model_params_cache WHERE signature = :sig LIMIT 1;"
    try:
        df = pd.read_sql(text(q), con=engine, params={"sig": sig})
        if df.empty:
            return None
        return df.iloc[0]["params_json"]
    except Exception:
        return None


def _save_params_to_db_cache(sig: str, params: Dict[str, Any]) -> None:
    _ensure_params_cache_table()
    q = """
    INSERT INTO model_params_cache(signature, params_json, updated_at)
    VALUES (:sig, :params_json::jsonb, NOW())
    ON CONFLICT (signature)
    DO UPDATE SET params_json = EXCLUDED.params_json, updated_at = NOW();
    """
    with engine.begin() as conn:
        conn.execute(text(q), {"sig": sig, "params_json": json.dumps(params, ensure_ascii=False)})


def ensure_params_loaded_from_db_silent() -> None:
    if "computed_params" in st.session_state and isinstance(st.session_state.computed_params, dict):
        return
    sig = _get_db_signature()
    cached = _load_params_from_db_cache(sig)
    if isinstance(cached, dict) and cached:
        st.session_state.computed_params = cached
        return


def compute_params_from_db() -> Dict[str, Any]:
    calv = pd.read_sql(
        "SELECT reg, mother_reg, birth_date, sex, event_type, event_date FROM calvings_births_raw",
        con=engine,
    )
    ins = pd.read_sql(
        "SELECT reg, lact, dim_age, event_date, bull, result FROM inseminations_raw",
        con=engine,
    )
    dry = pd.read_sql(
        "SELECT reg, dim, event_date FROM dryoff_raw",
        con=engine,
    )
    disp = pd.read_sql(
        "SELECT reg, event_date, disposal_reason FROM disposals_raw",
        con=engine,
    )
    bulls = pd.read_sql(
        "SELECT bull_code, bull_type FROM bulls_raw",
        con=engine,
    )

    out: Dict[str, Any] = {}

    # normalize
    calv["event_date"] = pd.to_datetime(calv["event_date"], errors="coerce")
    calv["birth_date"] = pd.to_datetime(calv["birth_date"], errors="coerce")
    calv["event_type_norm"] = calv["event_type"].apply(_norm_event_type)
    calv["reg_s"] = calv["reg"].apply(_norm_id)
    calv["mother_reg_s"] = calv["mother_reg"].apply(_norm_id)
    calv["sex_norm"] = calv["sex"].apply(_norm_sex)

    ins["event_date"] = pd.to_datetime(ins["event_date"], errors="coerce")
    ins["result_norm"] = ins["result"].apply(_norm_result)
    ins["reg_s"] = ins["reg"].apply(_norm_id)
    ins["bull_s"] = ins["bull"].apply(_norm_id)
    ins["lact"] = pd.to_numeric(ins["lact"], errors="coerce").fillna(0).astype(int)
    ins["dim_age"] = pd.to_numeric(ins["dim_age"], errors="coerce")

    dry["event_date"] = pd.to_datetime(dry["event_date"], errors="coerce")
    dry["reg_s"] = dry["reg"].apply(_norm_id)

    disp["event_date"] = pd.to_datetime(disp["event_date"], errors="coerce")
    disp["reg_s"] = disp["reg"].apply(_norm_id)
    disp["reason_norm"] = disp["disposal_reason"].astype(str).str.lower().str.replace("ё", "е")

    bulls["bull_code_s"] = bulls["bull_code"].apply(_norm_id)
    bulls["semen"] = bulls["bull_type"].apply(_classify_semen_from_bull_type)
    semen_by_bull = dict(zip(bulls["bull_code_s"].astype(str), bulls["semen"].astype(str)))

    # ------------------ conception params (P)
    ins_p = ins[(ins["event_date"].notna()) & (ins["result_norm"] == "P") & (ins["reg_s"] != "")].copy()

    avg_by_lact = dict(mp.CONCEPTION_PARAMS.avg_cow_dim_by_lact)
    avg_global = float(mp.CONCEPTION_PARAMS.avg_cow_dim_global)
    avg_heif_age = float(mp.CONCEPTION_PARAMS.avg_heifer_age_days)

    if not ins_p.empty:
        cows_p = ins_p[(ins_p["lact"] > 0) & ins_p["dim_age"].notna()].copy()
        if not cows_p.empty:
            cows_p["lact_cat"] = cows_p["lact"].clip(lower=1, upper=4)
            g = cows_p.groupby("lact_cat", sort=False)["dim_age"].mean().to_dict()
            avg_by_lact = {int(k): float(v) for k, v in g.items() if pd.notna(v)}
            gmean = cows_p["dim_age"].mean()
            if pd.notna(gmean):
                avg_global = float(gmean)

        heif_p = ins_p[(ins_p["lact"] <= 0) & (ins_p["dim_age"].notna())].copy()
        if not heif_p.empty:
            hmean = heif_p["dim_age"].mean()
            if pd.notna(hmean):
                avg_heif_age = float(hmean)

    out["conception"] = {
        "avg_cow_dim_by_lact": {
            1: float(avg_by_lact.get(1, avg_global)),
            2: float(avg_by_lact.get(2, avg_global)),
            3: float(avg_by_lact.get(3, avg_global)),
            4: float(avg_by_lact.get(4, avg_global)),
        },
        "avg_cow_dim_global": float(avg_global),
        "avg_heifer_age_days": float(avg_heif_age),
    }

    # ------------------ gestation days
    out["gestation_days"] = float(getattr(mp, "GESTATION_DAYS", 272.0))
    calves = calv[
        (calv["event_date"].notna())
        & (calv["event_type_norm"] == "РОЖДЕН")
        & (calv["mother_reg_s"] != "")
    ][["mother_reg_s", "event_date"]].drop_duplicates()

    if not calves.empty and not ins_p.empty:
        left = calves.rename(columns={"mother_reg_s": "reg_s", "event_date": "calving_dt"}).copy()
        right = ins_p.rename(columns={"event_date": "ins_dt"}).copy()
        right = right[["reg_s", "ins_dt"]].copy()

        pairs = _merge_asof_safe(
            left,
            right,
            by="reg_s",
            left_on="calving_dt",
            right_on="ins_dt",
            direction="backward",
            allow_exact_matches=True,
        )
        pairs = pairs[pairs["ins_dt"].notna()].copy()
        if not pairs.empty:
            pairs["gest_days"] = (pairs["calving_dt"] - pairs["ins_dt"]).dt.days
            pairs = pairs[(pairs["gest_days"] >= 200) & (pairs["gest_days"] <= 310)].copy()
            if not pairs.empty:
                out["gestation_days"] = float(pairs["gest_days"].mean())

    # ------------------ dry days
    out["dry_days"] = int(getattr(mp, "DRY_DAYS", 53))
    if not dry.empty and not calves.empty:
        left = calves.rename(columns={"mother_reg_s": "reg_s", "event_date": "calving_dt"}).copy()
        right = dry.rename(columns={"event_date": "dry_dt"}).copy()
        right = right[(right["dry_dt"].notna()) & (right["reg_s"] != "")][["reg_s", "dry_dt"]].copy()

        m = _merge_asof_safe(
            left,
            right,
            by="reg_s",
            left_on="calving_dt",
            right_on="dry_dt",
            direction="backward",
            allow_exact_matches=True,
        )
        m = m[m["dry_dt"].notna()].copy()
        if not m.empty:
            m["dry_days"] = (m["calving_dt"] - m["dry_dt"]).dt.days
            m = m[(m["dry_days"] >= 10) & (m["dry_days"] <= 200)].copy()
            if not m.empty:
                out["dry_days"] = int(round(float(m["dry_days"].median())))

    # ------------------ disposal params
    disposal_params = {
        "by_lact": {
            1: {"n": 0, "mean_dim": 0.0, "median_dim": 0.0},
            2: {"n": 0, "mean_dim": 0.0, "median_dim": 0.0},
            3: {"n": 0, "mean_dim": 0.0, "median_dim": 0.0},
            4: {"n": 0, "mean_dim": 0.0, "median_dim": 0.0},
        },
        "overall": {"n": 0, "mean_dim": 0.0, "median_dim": 0.0},
    }

    disp2 = disp[
        (disp["event_date"].notna())
        & (disp["reg_s"] != "")
        & (~disp["reason_norm"].str.contains("переезд", na=False))
    ].copy()

    ins2 = ins[(ins["event_date"].notna()) & (ins["reg_s"] != "") & ins["dim_age"].notna()].copy()
    ins2 = ins2[(ins2["dim_age"] >= 0) & (ins2["dim_age"] <= 500)].copy()

    if not disp2.empty and not ins2.empty:
        left = disp2.rename(columns={"event_date": "disp_dt"})[["reg_s", "disp_dt"]].copy()
        right = ins2.rename(columns={"event_date": "ins_dt"})[["reg_s", "ins_dt", "lact", "dim_age"]].copy()

        m = _merge_asof_safe(
            left,
            right,
            by="reg_s",
            left_on="disp_dt",
            right_on="ins_dt",
            direction="backward",
            allow_exact_matches=True,
        )
        m = m[m["ins_dt"].notna()].copy()
        if not m.empty:
            m["lact_cat"] = m["lact"].clip(lower=1, upper=4).astype(int)

            disposal_params["overall"]["n"] = int(len(m))
            disposal_params["overall"]["mean_dim"] = float(m["dim_age"].mean())
            disposal_params["overall"]["median_dim"] = float(m["dim_age"].median())

            for lc in (1, 2, 3, 4):
                sub = m[m["lact_cat"] == lc]
                if sub.empty:
                    continue
                disposal_params["by_lact"][lc]["n"] = int(len(sub))
                disposal_params["by_lact"][lc]["mean_dim"] = float(sub["dim_age"].mean())
                disposal_params["by_lact"][lc]["median_dim"] = float(sub["dim_age"].median())

    out["disposal_params"] = disposal_params
    out["annual_disposal_rate"] = float(getattr(mp, "ANNUAL_DISPOSAL_RATE", 0.0957))

    # ------------------ insemination params
    ip = {
        "cow_services_per_conception": float(mp.INSEMINATION_PARAMS.cow_services_per_conception),
        "cow_ai_interval_days": float(mp.INSEMINATION_PARAMS.cow_ai_interval_days),
        "cow_first_ai_dim_by_lact": dict(mp.INSEMINATION_PARAMS.cow_first_ai_dim_by_lact),
        "heifer_services_per_conception": float(mp.INSEMINATION_PARAMS.heifer_services_per_conception),
        "heifer_ai_interval_days": float(mp.INSEMINATION_PARAMS.heifer_ai_interval_days),
        "heifer_first_ai_age_days": float(mp.INSEMINATION_PARAMS.heifer_first_ai_age_days),
    }

    ins_all = ins[(ins["event_date"].notna()) & (ins["reg_s"] != "")].copy()
    if not ins_all.empty:
        # cows
        cows = ins_all[ins_all["lact"] > 0].copy()
        cows = cows[(cows["dim_age"].notna()) & (cows["dim_age"] >= 0) & (cows["dim_age"] <= 500)].copy()
        if not cows.empty:
            cows = cows.sort_values(["reg_s", "lact", "event_date"], kind="mergesort")
            grp = cows.groupby(["reg_s", "lact"], sort=False)

            services_counts: List[float] = []
            first_dims_by_lact: Dict[int, List[float]] = {1: [], 2: [], 3: [], 4: []}
            intervals: List[float] = []

            for (_reg, lact), g in grp:
                g = g.sort_values("event_date", kind="mergesort")
                mask_p = (g["result_norm"] == "P").to_numpy()
                if mask_p.sum() == 0:
                    continue
                first_p_pos = int(mask_p.argmax())
                services_counts.append(float(first_p_pos + 1))

                first_dim = g.iloc[0]["dim_age"]
                if pd.notna(first_dim):
                    lc = int(min(4, max(1, int(lact))))
                    first_dims_by_lact[lc].append(float(first_dim))

                dts = pd.to_datetime(g["event_date"]).sort_values()
                if len(dts) >= 2:
                    diffs = dts.diff().dt.days.dropna()
                    intervals.extend([float(x) for x in diffs.values if pd.notna(x)])

            if services_counts:
                ip["cow_services_per_conception"] = float(pd.Series(services_counts).mean())
            if intervals:
                ip["cow_ai_interval_days"] = float(pd.Series(intervals).mean())
            for lc in (1, 2, 3, 4):
                if first_dims_by_lact[lc]:
                    ip["cow_first_ai_dim_by_lact"][lc] = float(pd.Series(first_dims_by_lact[lc]).mean())

        # heifers
        heif = ins_all[ins_all["lact"] <= 0].copy()
        heif = heif[(heif["dim_age"].notna()) & (heif["dim_age"] >= 0) & (heif["dim_age"] <= 900)].copy()
        if not heif.empty:
            heif = heif.sort_values(["reg_s", "event_date"], kind="mergesort")
            grp = heif.groupby("reg_s", sort=False)

            services_counts_h: List[float] = []
            intervals_h: List[float] = []
            first_ages: List[float] = []

            for _reg, g in grp:
                g = g.sort_values("event_date", kind="mergesort")
                mask_p = (g["result_norm"] == "P").to_numpy()
                if mask_p.sum() == 0:
                    continue
                first_p_pos = int(mask_p.argmax())
                services_counts_h.append(float(first_p_pos + 1))

                first_age = g.iloc[0]["dim_age"]
                if pd.notna(first_age):
                    first_ages.append(float(first_age))

                dts = pd.to_datetime(g["event_date"]).sort_values()
                if len(dts) >= 2:
                    diffs = dts.diff().dt.days.dropna()
                    intervals_h.extend([float(x) for x in diffs.values if pd.notna(x)])

            if services_counts_h:
                ip["heifer_services_per_conception"] = float(pd.Series(services_counts_h).mean())
            if intervals_h:
                ip["heifer_ai_interval_days"] = float(pd.Series(intervals_h).mean())
            if first_ages:
                ip["heifer_first_ai_age_days"] = float(pd.Series(first_ages).mean())

    out["insemination_params"] = ip

    # ------------------ semen usage
    semen_usage = {
        "cow_trad": float(mp.SEMEN_USAGE_PROBS.cow_trad),
        "cow_sex": float(mp.SEMEN_USAGE_PROBS.cow_sex),
        "heifer_trad": float(mp.SEMEN_USAGE_PROBS.heifer_trad),
        "heifer_sex": float(mp.SEMEN_USAGE_PROBS.heifer_sex),
        "meta": {"window": "fallback", "n_cow": 0, "n_heifer": 0},
    }

    if not ins_p.empty:
        p = ins_p.copy()
        p["bull_s"] = p["bull_s"].astype(str)
        p["semen"] = p["bull_s"].map(semen_by_bull).fillna("trad")

        max_dt = p["event_date"].max()
        p_last = p
        window = "all"
        if pd.notna(max_dt):
            p_last = p[p["event_date"] >= (max_dt - pd.Timedelta(days=365))].copy()
            window = "last_year"

        def _shares(df: pd.DataFrame) -> Tuple[int, int, int]:
            if df.empty:
                return 0, 0, 0
            total = int(len(df))
            sex = int((df["semen"] == "sex").sum())
            trad = total - sex
            return trad, sex, total

        cow_df = p_last[p_last["lact"] > 0]
        heif_df = p_last[p_last["lact"] <= 0]

        cow_trad, cow_sex, n_cow = _shares(cow_df)
        hef_trad, hef_sex, n_hef = _shares(heif_df)

        if n_cow < 200 or n_hef < 200:
            window = "all"
            cow_trad, cow_sex, n_cow = _shares(p[p["lact"] > 0])
            hef_trad, hef_sex, n_hef = _shares(p[p["lact"] <= 0])

        prior_w = 500.0
        cow_sex_p = _bayes_smooth_share(cow_sex, n_cow, float(mp.SEMEN_USAGE_PROBS.cow_sex), prior_w)
        hef_sex_p = _bayes_smooth_share(hef_sex, n_hef, float(mp.SEMEN_USAGE_PROBS.heifer_sex), prior_w)

        semen_usage = {
            "cow_trad": float(1.0 - cow_sex_p),
            "cow_sex": float(cow_sex_p),
            "heifer_trad": float(1.0 - hef_sex_p),
            "heifer_sex": float(hef_sex_p),
            "meta": {"window": window, "n_cow": int(n_cow), "n_heifer": int(n_hef)},
        }

    out["semen_usage"] = semen_usage

    # ------------------ semen sex ratios
    semen_sex_ratios = {
        "trad": {
            "bull_share": float(mp.SEMEN_SEX_RATIOS["trad"].bull_share),
            "heifer_share": float(mp.SEMEN_SEX_RATIOS["trad"].heifer_share),
        },
        "sex": {
            "bull_share": float(mp.SEMEN_SEX_RATIOS["sex"].bull_share),
            "heifer_share": float(mp.SEMEN_SEX_RATIOS["sex"].heifer_share),
        },
        "meta": {"n_trad": 0, "n_sex": 0, "window": "fallback"},
    }

    calves_sex = calv[
        (calv["event_date"].notna())
        & (calv["event_type_norm"] == "РОЖДЕН")
        & (calv["mother_reg_s"] != "")
        & (calv["sex_norm"].isin(["F", "M"]))
    ][["mother_reg_s", "event_date", "sex_norm"]].copy()

    if not calves_sex.empty and not ins_p.empty:
        left = calves_sex.rename(columns={"mother_reg_s": "reg_s", "event_date": "calving_dt"}).copy()
        right = ins_p.rename(columns={"event_date": "ins_dt"}).copy()
        right = right[["reg_s", "ins_dt", "bull_s"]].copy()

        pairs = _merge_asof_safe(
            left,
            right,
            by="reg_s",
            left_on="calving_dt",
            right_on="ins_dt",
            direction="backward",
            allow_exact_matches=True,
        )
        pairs = pairs[pairs["ins_dt"].notna()].copy()
        if not pairs.empty:
            pairs["gest_days"] = (pairs["calving_dt"] - pairs["ins_dt"]).dt.days
            pairs = pairs[(pairs["gest_days"] >= 200) & (pairs["gest_days"] <= 310)].copy()

        if not pairs.empty:
            pairs["semen"] = pairs["bull_s"].astype(str).map(semen_by_bull).fillna("trad")

            max_dt = pairs["calving_dt"].max()
            window = "all"
            pairs_w = pairs
            if pd.notna(max_dt):
                pairs_w = pairs[pairs["calving_dt"] >= (max_dt - pd.Timedelta(days=365))].copy()
                window = "last_year"

            def _ratio(df: pd.DataFrame) -> Tuple[float, int]:
                if df.empty:
                    return 0.0, 0
                n = int(len(df))
                bulls_cnt = int((df["sex_norm"] == "M").sum())
                return float(bulls_cnt / max(1, n)), n

            trad_bull_p, n_trad = _ratio(pairs_w[pairs_w["semen"] == "trad"])
            sex_bull_p, n_sex = _ratio(pairs_w[pairs_w["semen"] == "sex"])

            if n_sex < 200:
                window = "all"
                trad_bull_p, n_trad = _ratio(pairs[pairs["semen"] == "trad"])
                sex_bull_p, n_sex = _ratio(pairs[pairs["semen"] == "sex"])

            prior_w = 300.0
            trad_bull_prior = float(mp.SEMEN_SEX_RATIOS["trad"].bull_share)
            sex_bull_prior = float(mp.SEMEN_SEX_RATIOS["sex"].bull_share)

            trad_bull_p = _bayes_smooth_share(trad_bull_p * n_trad, n_trad, trad_bull_prior, prior_w)
            sex_bull_p = _bayes_smooth_share(sex_bull_p * n_sex, n_sex, sex_bull_prior, prior_w)

            semen_sex_ratios = {
                "trad": {"bull_share": float(trad_bull_p), "heifer_share": float(1.0 - trad_bull_p)},
                "sex": {"bull_share": float(sex_bull_p), "heifer_share": float(1.0 - sex_bull_p)},
                "meta": {"n_trad": int(n_trad), "n_sex": int(n_sex), "window": window},
            }

    out["semen_sex_ratios"] = semen_sex_ratios

    return out


@st.cache_data(show_spinner=False)
def _compute_params_cached(db_signature: str) -> Dict[str, Any]:
    return compute_params_from_db()


def get_param_source() -> Dict[str, Any]:
    if "computed_params" in st.session_state and isinstance(st.session_state.computed_params, dict):
        return st.session_state.computed_params

    return {
        "conception": {
            "avg_cow_dim_by_lact": dict(mp.CONCEPTION_PARAMS.avg_cow_dim_by_lact),
            "avg_cow_dim_global": float(mp.CONCEPTION_PARAMS.avg_cow_dim_global),
            "avg_heifer_age_days": float(mp.CONCEPTION_PARAMS.avg_heifer_age_days),
        },
        "gestation_days": float(mp.GESTATION_DAYS),
        "dry_days": int(getattr(mp, "DRY_DAYS", 53)),
        "disposal_params": dict(mp.DISPOSAL_PARAMS),
        "annual_disposal_rate": float(getattr(mp, "ANNUAL_DISPOSAL_RATE", 0.0957)),
        "insemination_params": {
            "cow_services_per_conception": float(mp.INSEMINATION_PARAMS.cow_services_per_conception),
            "cow_ai_interval_days": float(mp.INSEMINATION_PARAMS.cow_ai_interval_days),
            "cow_first_ai_dim_by_lact": dict(mp.INSEMINATION_PARAMS.cow_first_ai_dim_by_lact),
            "heifer_services_per_conception": float(mp.INSEMINATION_PARAMS.heifer_services_per_conception),
            "heifer_ai_interval_days": float(mp.INSEMINATION_PARAMS.heifer_ai_interval_days),
            "heifer_first_ai_age_days": float(mp.INSEMINATION_PARAMS.heifer_first_ai_age_days),
        },
        "semen_usage": {
            "cow_trad": float(mp.SEMEN_USAGE_PROBS.cow_trad),
            "cow_sex": float(mp.SEMEN_USAGE_PROBS.cow_sex),
            "heifer_trad": float(mp.SEMEN_USAGE_PROBS.heifer_trad),
            "heifer_sex": float(mp.SEMEN_USAGE_PROBS.heifer_sex),
            "meta": {"window": "fallback", "n_cow": 0, "n_heifer": 0},
        },
        "semen_sex_ratios": {
            "trad": {
                "bull_share": float(mp.SEMEN_SEX_RATIOS["trad"].bull_share),
                "heifer_share": float(mp.SEMEN_SEX_RATIOS["trad"].heifer_share),
            },
            "sex": {
                "bull_share": float(mp.SEMEN_SEX_RATIOS["sex"].bull_share),
                "heifer_share": float(mp.SEMEN_SEX_RATIOS["sex"].heifer_share),
            },
            "meta": {"n_trad": 0, "n_sex": 0, "window": "fallback"},
        },
    }


def _apply_admin_overrides(base_params: Dict[str, Any]) -> Dict[str, Any]:
    p = {k: v for k, v in base_params.items()}
    ov = st.session_state.get("runtime_overrides")
    if not isinstance(ov, dict) or not ov:
        return p

    for key in ("gestation_days", "dry_days", "annual_disposal_rate"):
        if key in ov:
            p[key] = ov[key]

    if isinstance(ov.get("conception"), dict):
        p["conception"] = ov["conception"]
    if isinstance(ov.get("insemination_params"), dict):
        p["insemination_params"] = ov["insemination_params"]
    if isinstance(ov.get("semen_usage"), dict):
        p["semen_usage"] = ov["semen_usage"]
    if isinstance(ov.get("semen_sex_ratios"), dict):
        p["semen_sex_ratios"] = ov["semen_sex_ratios"]

    return p


# ============================================================
# UI
# ============================================================
st.set_page_config(page_title="Прогноз поголовья", layout="wide")
st.title("Прогноз поголовья по подразделению")

# чтобы в первом запуске не тормозило
ensure_params_loaded_from_db_silent()

# ---------- admin (one place, not sidebar)
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

ADMIN_KEY_TRUE = os.getenv("ADMIN_KEY", "admin")

with st.expander("Админ-режим", expanded=st.session_state.is_admin):
    if not st.session_state.is_admin:
        k = st.text_input("Ключ доступа", type="password", key="admin_key_input")
        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("Войти", key="admin_login_btn", use_container_width=True):
                if k == ADMIN_KEY_TRUE:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    st.error("Неверный ключ")
        with c2:
            st.caption("Ключ берётся из переменной окружения ADMIN_KEY (docker-compose).")
    else:
        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("Выйти", key="admin_logout_btn", use_container_width=True):
                st.session_state.is_admin = False
                st.session_state.pop("runtime_overrides", None)
                st.rerun()
        with c2:
            st.success("Админ-режим включён")

with st.expander("Справка: как считается прогноз", expanded=False):
    st.markdown(
        """
Мы строим **состояние стада** на дату последнего события в данных (максимальная дата среди: отёлы/рождения, осеменения, запуски, выбытия).

Дальше выполняем **симуляцию по дням**:
- животные “стареют” (DIM/возраст +1 в день),
- появляются новые стельности через параметры осеменения (первая попытка, интервал между попытками, доз на 1 стельность),
- стельные автоматически переходят в сухостой за число дней “длительность сухостоя” до отёла,
- происходят отёлы, добавляются телята по полу (по типу семени),
- происходит выбытие (по годовому проценту и форме по DIM/лактациям).

В прогнозе показывается **состояние на конец каждого месяца** и **ожидаемые отёлы в этом месяце**.
        """
    )

tab1, tab2, tab3 = st.tabs(
    [
        "Прогноз (одно подразделение)",
        "Параметры",
        "Прогноз по хозяйству (beta)",
    ]
)

# ============================================================
# TAB 2: параметры + админ-панель (вернул как было)
# ============================================================
with tab2:
    st.subheader("Параметры модели")

    cA, cB, cC = st.columns([1, 1, 2])
    with cA:
        if st.button("Пересчитать параметры из БД", use_container_width=True, key="btn_recalc_params_db"):
            try:
                _compute_params_cached.clear()
                sig = _get_db_signature()
                params = _compute_params_cached(sig)
                st.session_state.computed_params = params
                _save_params_to_db_cache(sig, params)
                st.success("Параметры пересчитаны и сохранены в кэш БД.")
            except Exception as e:
                st.error(f"Не удалось пересчитать параметры из данных: {e}")
                st.stop()

    with cB:
        if st.button("Сбросить кэш", use_container_width=True, key="btn_clear_cache"):
            st.cache_data.clear()
            st.success("Кэш очищен.")

    base_params = get_param_source()
    final_params_for_forecast = _apply_admin_overrides(base_params)

    with st.expander("Параметры модели (из данных/дефолтов)", expanded=False):
        st.markdown("### Вместимость (места)")
        st.table(pd.DataFrame([{"Группа": k, "Мест": int(v)} for k, v in mp.HERD_CAPACITY.items()]))

        conc = base_params.get("conception", {}) or {}
        st.markdown("### Стельность (по плодотворным осеменениям)")
        rows_p = []
        by_l = conc.get("avg_cow_dim_by_lact", {}) or {}
        for lact_cat in (1, 2, 3, 4):
            v = by_l.get(lact_cat)
            rows_p.append(
                {
                    "Лактация": {1: "1-я", 2: "2-я", 3: "3-я", 4: "4+ (и старше)"}[lact_cat],
                    "Средний DIM стельности, дни": round(float(v), 1) if v is not None else None,
                }
            )
        st.table(pd.DataFrame(rows_p))
        st.write(f"Средний DIM стельности по коровам (в целом): {float(conc.get('avg_cow_dim_global', 0.0)):.1f} дней")
        st.write(
            f"Средний возраст стельности тёлок: {float(conc.get('avg_heifer_age_days', 0.0)):.1f} дней "
            f"(≈ {float(conc.get('avg_heifer_age_days', 0.0)) / 30:.1f} месяцев)"
        )

        st.markdown("### Длительность стельности и сухостоя")
        st.write(f"Длительность стельности: {float(base_params.get('gestation_days', 0.0)):.1f} дней")
        st.write(f"Длительность сухостоя: {int(base_params.get('dry_days', 0) or 0)} дней")

        st.markdown("### Выбытие коров (форма по DIM и лактациям)")
        st.write(f"Годовой процент выбытия (в модели): {float(base_params.get('annual_disposal_rate', 0.0)) * 100:.2f} %")

        disp_p = base_params.get("disposal_params", {}) or {}
        by_l2 = (disp_p.get("by_lact", {}) or {})
        overall_n = float((disp_p.get("overall", {}) or {}).get("n", 0) or 0)
        disp_rows = []
        for lact_cat in (1, 2, 3, 4):
            s = by_l2.get(lact_cat, {}) or {}
            n = float(s.get("n", 0) or 0)
            share = (n / overall_n * 100.0) if overall_n > 0 else 0.0
            disp_rows.append(
                {
                    "Лактация": {1: "1-я", 2: "2-я", 3: "3-я", 4: "4+ (и старше)"}[lact_cat],
                    "Доля выбытий среди выбывших, %": round(share, 2),
                    "Медианный DIM выбытия, дни": round(float(s.get("median_dim", 0.0) or 0.0), 1),
                    "Средний DIM выбытия, дни": round(float(s.get("mean_dim", 0.0) or 0.0), 1),
                }
            )
        st.table(pd.DataFrame(disp_rows))
        st.caption("DIM = число дней после отёла (Days In Milk) на момент события выбытия.")

        st.markdown("### Осеменения (как часто и сколько доз)")
        ins_p = base_params.get("insemination_params", {}) or {}
        st.table(
            pd.DataFrame(
                [
                    {"Показатель": "Коровы: доз на 1 стельность", "Значение": round(float(ins_p.get("cow_services_per_conception", 0.0)), 3)},
                    {"Показатель": "Коровы: интервал между осеменениями, дни", "Значение": round(float(ins_p.get("cow_ai_interval_days", 0.0)), 3)},
                    {"Показатель": "Тёлки: доз на 1 стельность", "Значение": round(float(ins_p.get("heifer_services_per_conception", 0.0)), 3)},
                    {"Показатель": "Тёлки: интервал между осеменениями, дни", "Значение": round(float(ins_p.get("heifer_ai_interval_days", 0.0)), 3)},
                    {"Показатель": "Тёлки: возраст первого осеменения, дни", "Значение": round(float(ins_p.get("heifer_first_ai_age_days", 0.0)), 3)},
                ]
            )
        )

    # ---------- Admin parameter editor (оставил как было по логике)
    if st.session_state.is_admin:
        with st.expander("Админ-панель: ручная настройка параметров", expanded=True):
            curr = final_params_for_forecast

            c = curr.get("conception", {}) or {}
            ip = curr.get("insemination_params", {}) or {}
            su = curr.get("semen_usage", {}) or {}
            ssr = curr.get("semen_sex_ratios", {}) or {}

            with st.form("admin_params_form"):
                st.markdown("### Длительности")
                gest = st.number_input(
                    "Длительность стельности (дни)",
                    min_value=200.0,
                    max_value=310.0,
                    value=float(curr.get("gestation_days", 272.0)),
                    step=1.0,
                    key="adm_gest",
                )
                dryd = st.number_input(
                    "Длительность сухостоя (дни)",
                    min_value=20,
                    max_value=150,
                    value=int(curr.get("dry_days", 53)),
                    step=1,
                    key="adm_dry",
                )

                st.markdown("### Стельность (целевые средние)")
                colA, colB, colC, colD = st.columns(4)
                l1 = float(colA.number_input("Коровы: средний DIM стельности, 1-я лактация", value=float(c.get("avg_cow_dim_by_lact", {}).get(1, 99.0)), step=0.1, key="adm_c_l1"))
                l2 = float(colB.number_input("Коровы: средний DIM стельности, 2-я лактация", value=float(c.get("avg_cow_dim_by_lact", {}).get(2, 107.0)), step=0.1, key="adm_c_l2"))
                l3 = float(colC.number_input("Коровы: средний DIM стельности, 3-я лактация", value=float(c.get("avg_cow_dim_by_lact", {}).get(3, 105.0)), step=0.1, key="adm_c_l3"))
                l4 = float(colD.number_input("Коровы: средний DIM стельности, 4+ лактация", value=float(c.get("avg_cow_dim_by_lact", {}).get(4, 107.0)), step=0.1, key="adm_c_l4"))
                cg = float(st.number_input("Коровы: средний DIM стельности (в целом)", value=float(c.get("avg_cow_dim_global", 104.0)), step=0.1, key="adm_c_g"))
                ha = float(st.number_input("Тёлки: средний возраст стельности (дни)", value=float(c.get("avg_heifer_age_days", 402.0)), step=0.1, key="adm_h_a"))

                st.markdown("### Выбытие")
                adr = float(
                    st.number_input(
                        "Годовой процент выбытия (доля в год)",
                        min_value=0.0,
                        max_value=0.5,
                        value=float(curr.get("annual_disposal_rate", 0.0957)),
                        step=0.001,
                        format="%.3f",
                        key="adm_adr",
                    )
                )

                st.markdown("### Осеменения")
                cow_spc = float(st.number_input("Коровы: доз на 1 стельность", min_value=1.0, max_value=6.0, value=float(ip.get("cow_services_per_conception", 2.04)), step=0.01, key="adm_c_spc"))
                cow_int = float(st.number_input("Коровы: интервал между осеменениями (дни)", min_value=10.0, max_value=120.0, value=float(ip.get("cow_ai_interval_days", 46.8)), step=0.5, key="adm_c_int"))
                heif_spc = float(st.number_input("Тёлки: доз на 1 стельность", min_value=1.0, max_value=6.0, value=float(ip.get("heifer_services_per_conception", 1.95)), step=0.01, key="adm_h_spc"))
                heif_int = float(st.number_input("Тёлки: интервал между осеменениями (дни)", min_value=10.0, max_value=120.0, value=float(ip.get("heifer_ai_interval_days", 25.3)), step=0.5, key="adm_h_int"))
                heif_first = float(st.number_input("Тёлки: возраст первого осеменения (дни)", min_value=200.0, max_value=900.0, value=float(ip.get("heifer_first_ai_age_days", 378.5)), step=1.0, key="adm_h_first"))

                st.markdown("DIM первого осеменения по лактациям")
                cf = dict(ip.get("cow_first_ai_dim_by_lact", {1: 71.4, 2: 72.2, 3: 73.3, 4: 72.9}))
                cc1, cc2, cc3, cc4 = st.columns(4)
                cf1 = float(cc1.number_input("1-я лактация", value=float(cf.get(1, 71.4)), step=0.1, key="adm_cf1"))
                cf2 = float(cc2.number_input("2-я лактация", value=float(cf.get(2, 72.2)), step=0.1, key="adm_cf2"))
                cf3 = float(cc3.number_input("3-я лактация", value=float(cf.get(3, 73.3)), step=0.1, key="adm_cf3"))
                cf4 = float(cc4.number_input("4+ лактация", value=float(cf.get(4, 72.9)), step=0.1, key="adm_cf4"))

                st.markdown("### Доля использования семени (если хочешь вручную)")
                su_c1, su_c2, su_c3, su_c4 = st.columns(4)
                cow_trad = float(su_c1.number_input("Коровы: обычное (доля)", min_value=0.0, max_value=1.0, value=float(su.get("cow_trad", 0.7)), step=0.01, key="adm_su_ct"))
                cow_sex = float(su_c2.number_input("Коровы: сексированное (доля)", min_value=0.0, max_value=1.0, value=float(su.get("cow_sex", 0.3)), step=0.01, key="adm_su_cs"))
                heif_trad = float(su_c3.number_input("Тёлки: обычное (доля)", min_value=0.0, max_value=1.0, value=float(su.get("heifer_trad", 0.3)), step=0.01, key="adm_su_ht"))
                heif_sex = float(su_c4.number_input("Тёлки: сексированное (доля)", min_value=0.0, max_value=1.0, value=float(su.get("heifer_sex", 0.7)), step=0.01, key="adm_su_hs"))

                st.markdown("### Пол телят по типу семени (если хочешь вручную)")
                r1, r2 = st.columns(2)
                trad_bull = float(r1.number_input("Обычное: доля бычков", min_value=0.0, max_value=1.0, value=float(ssr.get("trad", {}).get("bull_share", 0.2483)), step=0.01, key="adm_trad_b"))
                sex_bull = float(r2.number_input("Сексированное: доля бычков", min_value=0.0, max_value=1.0, value=float(ssr.get("sex", {}).get("bull_share", 0.0583)), step=0.01, key="adm_sex_b"))

                saved = st.form_submit_button("Сохранить параметры для прогноза", use_container_width=True)
                if saved:
                    def _norm2(a: float, b: float) -> Tuple[float, float]:
                        s = max(1e-9, a + b)
                        return a / s, b / s

                    cow_trad, cow_sex = _norm2(cow_trad, cow_sex)
                    heif_trad, heif_sex = _norm2(heif_trad, heif_sex)

                    trad_bull = max(0.0, min(1.0, trad_bull))
                    sex_bull = max(0.0, min(1.0, sex_bull))

                    st.session_state.runtime_overrides = {
                        "gestation_days": float(gest),
                        "dry_days": int(dryd),
                        "annual_disposal_rate": float(adr),
                        "conception": {
                            "avg_cow_dim_by_lact": {1: l1, 2: l2, 3: l3, 4: l4},
                            "avg_cow_dim_global": cg,
                            "avg_heifer_age_days": ha,
                        },
                        "insemination_params": {
                            "cow_services_per_conception": cow_spc,
                            "cow_ai_interval_days": cow_int,
                            "cow_first_ai_dim_by_lact": {1: cf1, 2: cf2, 3: cf3, 4: cf4},
                            "heifer_services_per_conception": heif_spc,
                            "heifer_ai_interval_days": heif_int,
                            "heifer_first_ai_age_days": heif_first,
                        },
                        "semen_usage": {
                            "cow_trad": cow_trad,
                            "cow_sex": cow_sex,
                            "heifer_trad": heif_trad,
                            "heifer_sex": heif_sex,
                            "meta": {"window": "admin", "n_cow": 0, "n_heifer": 0},
                        },
                        "semen_sex_ratios": {
                            "trad": {"bull_share": trad_bull, "heifer_share": 1.0 - trad_bull},
                            "sex": {"bull_share": sex_bull, "heifer_share": 1.0 - sex_bull},
                            "meta": {"window": "admin", "n_trad": 0, "n_sex": 0},
                        },
                    }
                    st.success("Сохранено. Перейди на вкладку «Прогноз» и нажми «Рассчитать прогноз».")


# ============================================================
# TAB 3: хозяйство (beta) — вернул как было
# ============================================================
with tab3:
    st.subheader("Прогноз по хозяйству / несколько подразделений (beta)")

    st.info(
        "Эта вкладка пока с пустой логикой расчёта.\n\n"
        "Здесь будет: выбор нескольких подразделений и расчёт прогноза по ним, "
        "а также загрузка файла с данными по всему хозяйству."
    )

    uploaded = st.file_uploader(
        "Загрузить файл по всему хозяйству (xlsx/csv/zip) — пока только сохраняем и показываем метаданные",
        type=["xlsx", "xls", "csv", "zip"],
        accept_multiple_files=False,
        key="farm_file_uploader",
    )
    if uploaded is not None:
        st.session_state["farm_file_name"] = uploaded.name
        st.session_state["farm_file_bytes"] = uploaded.getvalue()

        st.success(f"Файл загружен: {uploaded.name} ({len(st.session_state['farm_file_bytes'])} bytes)")

        try:
            if uploaded.name.lower().endswith((".xlsx", ".xls")):
                xls = pd.ExcelFile(uploaded)
                st.write("Листы в Excel:", xls.sheet_names)
            elif uploaded.name.lower().endswith(".csv"):
                st.write("Первые строки CSV:")
                df_preview = pd.read_csv(uploaded, nrows=20)
                st.dataframe(df_preview, use_container_width=True)
            else:
                st.write("ZIP загружен. Распаковку/разбор добавим после того как узнаем структуру.")
        except Exception as e:
            st.warning(f"Не удалось прочитать содержимое для превью: {e}")

    st.divider()

    st.caption("Выбор подразделений (пока список пустой — подключим после уточнения структуры данных):")
    departments = st.multiselect("Подразделения", options=[], default=[], key="farm_departments_multiselect")

    st.button("Посчитать прогноз по хозяйству", disabled=True, key="farm_calc_btn_disabled")


# ============================================================
# TAB 1: прогноз (одно подразделение)
# - убрал таблицу "Реализация" (по твоей просьбе)
# - таблицы ориентированы: СТОЛБЦЫ = месяцы
# - блок загрузки файлов скрыт, если выбран режим "данные из БД"
# ============================================================
with tab1:
    st.subheader("Источник данных")

    data_mode = st.radio(
        "Откуда брать данные для расчёта?",
        options=[
            "Использовать данные из БД (по умолчанию)",
            "Обновить БД из файлов",
        ],
        index=0,
        horizontal=True,
        key="data_mode_radio",
    )
    need_files = data_mode.startswith("Обновить БД")

    if not need_files:
        st.info("Расчёт возьмёт данные из БД. Блок загрузки файлов скрыт (он нужен только для обновления БД).")
    else:
        st.warning("В этом режиме при расчёте данные в БД будут заменены загруженными файлами (replace).")

        st.subheader("Загрузка файлов (для обновления БД)")
        col1, col2 = st.columns(2)
        with col1:
            calvings_file = st.file_uploader("Отёлы + родившиеся", type=["xls", "xlsx"], key="u_calvings")
            disposals_file = st.file_uploader("Выбытие", type=["xls", "xlsx"], key="u_disposals")
        with col2:
            dryoff_file = st.file_uploader("Запуски", type=["xls", "xlsx"], key="u_dryoff")
            inseminations_file = st.file_uploader("Осеменения", type=["xls", "xlsx"], key="u_inseminations")

        bulls_file = st.file_uploader("Таблица быков (txt)", type=["txt"], key="u_bulls")
    # чтобы имена существовали дальше
    if not need_files:
        calvings_file = None
        disposals_file = None
        dryoff_file = None
        inseminations_file = None
        bulls_file = None

    st.subheader("Месяц прогноза")

    months_ru = [
        "01 — Январь", "02 — Февраль", "03 — Март", "04 — Апрель",
        "05 — Май", "06 — Июнь", "07 — Июль", "08 — Август",
        "09 — Сентябрь", "10 — Октябрь", "11 — Ноябрь", "12 — Декабрь",
    ]

    today = date.today()
    sel_col1, sel_col2 = st.columns([1, 1])
    with sel_col1:
        year_sel = st.number_input("Год", min_value=2000, max_value=2100, value=today.year, step=1, key="year_sel")
    with sel_col2:
        month_sel_label = st.selectbox("Месяц", months_ru, index=today.month - 1, key="month_sel_label")
        month_sel = int(month_sel_label.split("—")[0].strip())

    target_month_end = month_end(int(year_sel), int(month_sel))

    st.subheader("Расчёт")
    calculate = st.button("Рассчитать прогноз", key="btn_calc_forecast", use_container_width=True)

    st.session_state.setdefault("last_result_df", None)          # индекс = месяцы
    st.session_state.setdefault("last_overflow_df", None)        # индекс = месяцы
    st.session_state.setdefault("last_month_ends", None)
    st.session_state.setdefault("last_excel_bytes", None)

    if calculate:
        # 1) если нужно — обновляем БД из файлов
        if need_files:
            missing = []
            if calvings_file is None:
                missing.append("Отёлы + родившиеся")
            if disposals_file is None:
                missing.append("Выбытие")
            if dryoff_file is None:
                missing.append("Запуски")
            if inseminations_file is None:
                missing.append("Осеменения")

            if missing:
                st.error("Не все файлы загружены: " + ", ".join(missing))
                st.stop()

            try:
                calv_df = read_calvings_excel(calvings_file)
                load_calvings_to_db(calv_df, if_exists="replace")
            except Exception as e:
                st.error(f"Ошибка при загрузке 'Отёлы + родившиеся': {e}")
                st.stop()

            try:
                disp_df = read_disposals_excel(disposals_file)
                load_disposals_to_db(disp_df, if_exists="replace")
            except Exception as e:
                st.error(f"Ошибка при загрузке 'Выбытие': {e}")
                st.stop()

            try:
                dry_df = read_dryoff_excel(dryoff_file)
                load_dryoff_to_db(dry_df, if_exists="replace")
            except Exception as e:
                st.error(f"Ошибка при загрузке 'Запуски': {e}")
                st.stop()

            try:
                ins_df = read_inseminations_excel(inseminations_file)
                ins_df = clean_inseminations(ins_df)
                load_inseminations_to_db(ins_df, if_exists="replace")
            except Exception as e:
                st.error(f"Ошибка при загрузке 'Осеменения': {e}")
                st.stop()

            if bulls_file is not None:
                try:
                    bulls_df = read_bulls_txt(bulls_file)
                    load_bulls_to_db(bulls_df, if_exists="replace")
                except Exception as e:
                    st.error(f"Ошибка при разборе 'Таблица быков': {e}")
                    st.stop()

            # recompute params после обновления
            try:
                with st.spinner("Пересчитываю параметры из загруженных данных..."):
                    params = compute_params_from_db()
                    st.session_state.computed_params = params
                    _save_params_to_db_cache(_get_db_signature(), params)
            except Exception as e:
                st.error(f"Не удалось пересчитать параметры из данных: {e}")
                st.stop()

        # 2) период прогноза от последней даты в БД
        base_date = get_max_event_date_from_db()
        base_month_end = month_end(base_date.year, base_date.month)

        if target_month_end < base_month_end:
            month_ends = [target_month_end]
        else:
            month_ends = iter_month_ends(
                base_date.year,
                base_date.month,
                target_month_end.year,
                target_month_end.month,
            )

        st.markdown(f"**Период прогноза:** {month_ends[0].strftime('%m.%Y')} → {month_ends[-1].strftime('%m.%Y')}")

        base_params = get_param_source()
        final_params_for_forecast = _apply_admin_overrides(base_params)

        rows: list[dict] = []
        overflow_rows: list[dict] = []

        prog = st.progress(0.0)
        with st.spinner("Считаю прогноз..."):
            for i, d_end in enumerate(month_ends, start=1):
                try:
                    vals = compute_forecast_from_db(d_end, overrides=final_params_for_forecast) or {}
                except Exception as e:
                    st.error(f"Ошибка расчёта на {d_end.strftime('%Y-%m')}: {e}")
                    vals = {}

                norm_map = {_norm_label(k2): v2 for k2, v2 in (vals or {}).items()}

                row = {"Месяц": d_end.strftime("%Y-%m")}
                for k in INDICATORS:
                    row[k] = _vals_get(vals, k, norm_map)
                rows.append(row)

                ov_row = {"Месяц": d_end.strftime("%Y-%m")}
                for k in OVERFLOW_COLS:
                    v = _vals_get(vals, k, norm_map)
                    ov_row[k] = 0.0 if v is None else v
                overflow_rows.append(ov_row)

                prog.progress(i / max(1, len(month_ends)))
        prog.empty()

        result_df = ensure_month_col(
            pd.DataFrame(rows),
            month_labels=[r.get("Месяц", "") for r in rows],
        ).set_index("Месяц")

        overflow_df = ensure_month_col(
            pd.DataFrame(overflow_rows),
            month_labels=[r.get("Месяц", "") for r in overflow_rows],
        ).set_index("Месяц")

        st.session_state["last_result_df"] = result_df
        st.session_state["last_overflow_df"] = overflow_df
        st.session_state["last_month_ends"] = month_ends

    # ----------------------------
    # РЕЗУЛЬТАТЫ (месяцы = столбцы)
    # ----------------------------
    result = st.session_state.get("last_result_df")
    overflow_df = st.session_state.get("last_overflow_df")
    month_ends = st.session_state.get("last_month_ends")

    if not isinstance(result, pd.DataFrame) or result.empty:
        st.info("Нажми «Рассчитать прогноз», чтобы увидеть таблицы.")
    else:
        # mapping: индикатор -> колонка переполнения
        indicator_to_overflow = {
            "Дойные коровы": "Переполнение: Дойные коровы",
            "Сухостойные коровы": "Переполнение: Сухостойные коровы",
            "Тёлки 0–3 мес": "Переполнение: Тёлки 0–3 мес",
            "Бычки 0–2 мес": None,
            "Тёлки 3–8 мес": "Переполнение: Тёлки 3–8 мес",
            "Тёлки ≥9 мес": "Переполнение: Тёлки 9–24 мес",
            "Нетели": "Переполнение: Нетели",
            "Ожидаемый отёл, всего": None,
            "Ожидаемый отёл, из них коров": None,
            "Ожидаемый отёл, из них нетелей": None,
            "Ожидаемые бычки (условно)": None,
            "Ожидаемые тёлочки (условно)": None,
        }

        BAD = "background-color: #ff0000; color: #ffffff; font-weight: 700;"

        # ВИЗУАЛ: делаем месяцы столбцами
        forecast_view = result.T  # индикаторы x месяцы

        overflow_groups_only = overflow_df.reindex(columns=[c for c in OVERFLOW_GROUP_COLS if c in overflow_df.columns])
        overflow_view = overflow_groups_only.T  # переполнение-группы x месяцы

        def style_forecast_months_as_columns(df_view: pd.DataFrame) -> pd.DataFrame:
            styles = pd.DataFrame("", index=df_view.index, columns=df_view.columns)
            # df_view: index = индикаторы, columns = месяцы
            for ind in df_view.index:
                ov_name = indicator_to_overflow.get(str(ind))
                if not ov_name:
                    continue
                if ov_name not in overflow_df.columns:
                    continue
                for m in df_view.columns:
                    try:
                        ov = float(pd.to_numeric(overflow_df.loc[str(m), ov_name], errors="coerce") or 0.0)
                    except Exception:
                        ov = 0.0
                    if ov > 0.0:
                        styles.loc[ind, m] = BAD
            return styles

        def style_positive_red(df_any: pd.DataFrame) -> pd.DataFrame:
            s = pd.DataFrame("", index=df_any.index, columns=df_any.columns)
            num = df_any.apply(pd.to_numeric, errors="coerce").fillna(0.0)
            s[num > 0.0] = BAD
            return s

        st.subheader("Прогноз (месяцы — столбцы)")
        st.dataframe(
            forecast_view.style.format(_fmt_cell).apply(style_forecast_months_as_columns, axis=None),
            use_container_width=True,
        )

        st.caption(
            "Важно: в прогнозе показаны значения УЖЕ после ограничения по местам (лишние не учитываются). "
            "Насколько “лишних” — см. таблицу переполнения ниже."
        )

        st.subheader("Переполнение по группам (месяцы — столбцы)")
        st.dataframe(
            overflow_view.style.format(_fmt_cell).apply(style_positive_red, axis=None),
            use_container_width=True,
        )

        # ----------------------------
        # Скачать Excel
        # ----------------------------
        st.subheader("Скачать результат (Excel)")
        excel_bytes = make_excel_bytes_highlight_months_columns(
            forecast_view=forecast_view,
            overflow_view=overflow_view,
            indicator_to_overflow=indicator_to_overflow,
        )
        st.session_state["last_excel_bytes"] = excel_bytes

        if isinstance(month_ends, list) and month_ends:
            file_name = f"herd_forecast_{month_ends[0].strftime('%Y-%m')}_to_{month_ends[-1].strftime('%Y-%m')}.xlsx"
        else:
            file_name = "herd_forecast.xlsx"

        st.download_button(
            label="Скачать Excel: прогноз + переполнение (месяцы столбцами, с подсветкой)",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_excel",
        )


# ============================================================
# LEGACY / НЕ УДАЛЯЕМ — но не исполняем
# ============================================================
if False:
    pass
