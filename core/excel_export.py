from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd

def make_excel_bytes(forecast_df: pd.DataFrame, realization_df: pd.DataFrame) -> bytes:
    import openpyxl
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        forecast_df.to_excel(writer, sheet_name="Прогноз")
        realization_df.to_excel(writer, sheet_name="Реализация")

        wb = writer.book
        fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

        def _apply_positive_rule(ws) -> None:
            if ws.max_row < 2 or ws.max_column < 2:
                return
            header_row = 1
            for col_idx in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{header_row+1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator="greaterThan", formula=["0"], fill=fill),
                )

        if "Реализация" in wb.sheetnames:
            _apply_positive_rule(wb["Реализация"])

    return out.getvalue()

def make_excel_bytes_highlight_months_columns(
    forecast_view: pd.DataFrame,
    overflow_view: pd.DataFrame,
    indicator_to_overflow: dict[str, str | None],
    realization_view: pd.DataFrame | None = None,
) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        forecast_view.to_excel(writer, sheet_name="Прогноз ")
        overflow_view.to_excel(writer, sheet_name="Переполнение ")
        if realization_view is not None and isinstance(realization_view, pd.DataFrame) and not realization_view.empty:
            realization_view.to_excel(writer, sheet_name="План реализации ")

    buf.seek(0)
    wb = load_workbook(buf)

    fill = PatternFill("solid", fgColor="FF0000")
    font = Font(color="FFFFFF", bold=True)
    num_fmt = "0"

    def _is_pos(v: Any) -> bool:
        try:
            return float(v) > 0.0
        except Exception:
            return False

    ws_ov = wb["Переполнение "]
    for row in ws_ov.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if _is_pos(cell.value):
                cell.fill = fill
                cell.font = font
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt

    ws_f = wb["Прогноз "]
    months = [str(ws_f.cell(row=1, column=c).value) for c in range(2, ws_f.max_column + 1)]
    indicators = [str(ws_f.cell(row=r, column=1).value) for r in range(2, ws_f.max_row + 1)]

    ov_month_to_col = {str(ws_ov.cell(row=1, column=c).value): c for c in range(2, ws_ov.max_column + 1)}
    ov_row_by_name = {str(ws_ov.cell(row=r, column=1).value): r for r in range(2, ws_ov.max_row + 1)}

    for r_idx, ind in enumerate(indicators, start=2):
        ov_name = indicator_to_overflow.get(ind)
        if not ov_name:
            continue
        ov_r = ov_row_by_name.get(ov_name)
        if not ov_r:
            continue
        for c_idx, m in enumerate(months, start=2):
            ov_c = ov_month_to_col.get(m)
            if not ov_c:
                continue
            ov_val = ws_ov.cell(row=ov_r, column=ov_c).value
            cell = ws_f.cell(row=r_idx, column=c_idx)
            if _is_pos(ov_val):
                cell.fill = fill
                cell.font = font
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt

    if "План реализации " in wb.sheetnames:
        ws_r = wb["План реализации "]
        for row in ws_r.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if _is_pos(cell.value):
                    cell.fill = fill
                    cell.font = font
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
